import argparse
import timeit
import os
from datetime import date

import numpy as np
import pandas as pd
import pypyodbc
from openpyxl import load_workbook


def create_property_table_dummy(db_path, out_path, files_name, db_name):
    # TODO: Implement using another library.
    print('Implementation Pending')
    with open(f'{out_path}/{files_name}.csv', 'w') as fp:
        pass


def create_property_table(db_path, out_path, db_name, excel_name, files_name_suffix):
    create_excel_template(db_path, excel_name)
    conn = pypyodbc.connect(
        r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};" +
        f"Dbq={db_path}/{db_name};"
        )
    cur = conn.cursor()
    print(f"--- Creating GENERATION db_key to data dictionary of {db_name} ---")
    gen_key_to_data = get_key_to_data_dict(cur, type=("G"))  # "Line", "Flow";  "Node", "Genration"; "Node", "Marginal Loss Factor"
    print(f"--- Creating DEMAND db_key to name dictionary of {db_name} ---")
    dem_key_to_data = get_key_to_data_dict(cur, type=("L", "L_D", "R"))  # "Line", "Flow";  "Node", "Genration"; "Node", "Marginal Loss Factor"
    print(f"--- Creating GENERATION table of {db_name} ---")
    files_name = f"GENERATION_{files_name_suffix}"
    gen_table_wide = database_to_table(gen_key_to_data, cur, out_path, files_name)
    print(f"--- Creating DEMAND table of {db_name} ---")
    files_name = f"DEMAND_{files_name_suffix}"
    dem_table_wide = database_to_table(dem_key_to_data, cur, out_path, files_name)

    create_excel_template(db_path, excel_name)
    generate_days_excel(gen_key_to_data, dem_key_to_data, gen_table_wide, dem_table_wide, out_path, files_name_suffix)
    #print(f"Total of {class_name}s with {property_name} data = {property_table_wide.shape[0]}; Total of {class_name}s with No data = {len(key_to_class.keys()) - property_table_wide.shape[0]}")
    #print(f"--- {property_name} Table of {db_path[-66:-50]} created ---")
    #print("\n")
    cur.close()
    conn.close()


def get_key_to_data_dict(cur, type):
    # Get class_id
    cur.execute(f"SELECT clave, nombre_barra, Zona, tipo1 FROM BASE where tipo1 in{tuple(type)}")
    key_to_name = {}
    while True:
        row = cur.fetchone()
        if row is None:
            break
        key_to_name[row.get("clave")] = [row.get("tipo1"), row.get("nombre_barra"), row.get("Zona")]
        # print(u"{0} is class id {1}".format(row.get("name"), row.get("class_id")))

    return key_to_name


def database_to_table(key_to_data, cur, out_folder, files_name):
    t_data = "MEDIDAS"
    year = int(f'20{files_name[-4:-2]}')
    month = int(files_name[-2:])    
    if month < 12:
        num_days = (date(year, month + 1, 1) - date(year, month, 1)).days        
    elif month == 12:
        num_days = 31
    num_hours = num_days * 24
    dtype = [('clave_ds', 'U100'), ('hora_mensual', int), ('valor', float)]
    property_mx = np.zeros((len(key_to_data.keys()) * num_hours), dtype=dtype)

    # Get property values
    cur.execute(f"SELECT Hora_Mensual, MedidaHoraria, clave FROM {t_data} where clave in{tuple(key_to_data.keys())} and Hora_Mensual < {num_hours + 1}")
    mx_row = 0
    while True:
        row = cur.fetchone()
        if row is None:
            break
        property_mx[mx_row] = (row.get("clave"), row.get("Hora_Mensual"), row.get("MedidaHoraria"))
        mx_row += 1
    property_mx = np.sort(property_mx, order=['clave_ds', 'hora_mensual'])
    property_table = pd.DataFrame(property_mx)
    property_table = property_table.drop_duplicates()  # Drop all empty rows
    if property_table.iloc[0, 0] == '':
        property_table = property_table.drop(0)
    property_table_wide = pd.pivot(property_table, index="clave_ds", columns='hora_mensual', values='valor')
    # property_table_wide.to_csv(f"{out_folder}/{files_name}.csv", index=True, sep=";", decimal=",")
    
    return property_table_wide


def create_excel_template(path, excel_name):
    workbook = load_workbook(filename=f"{path}/{excel_name}")
    print(workbook.sheetnames)


    sheet = workbook.get_sheet_by_name('Inyecciones Fact')
    sheet["A1"] = "hello"
    sheet["B1"] = "world!"

    workbook.save(filename="caca.xlsx")



def generate_days_excel(gen_key_to_data, dem_key_to_data, gen_table_wide, dem_table_wide, out_folder, files_name_suffix):
    num_days = int(gen_table_wide.shape[1]/24)
    gen_data_table = pd.DataFrame.from_dict(gen_key_to_data, orient='index',
                       columns=['tipo1', 'nombre_barra', 'Zona'])
    gen_data_table = gen_data_table.sort_index() 
    gen_data_table['clave_ds'] = gen_data_table.index
    gen_data_table = gen_data_table.reset_index(drop=True)

              
    dem_data_table = pd.DataFrame.from_dict(dem_key_to_data, orient='index',
                       columns=['tipo1', 'nombre_barra', 'Zona'])
    dem_data_table = dem_data_table.sort_index()
    dem_data_table['clave_ds'] = dem_data_table.index
    dem_data_table = dem_data_table.reset_index(drop=True)

    # num_days = 2  # BORRAR
    for day in range(num_days):
        # TODO: Create Excel

        first_hour = day*24
        gen_day_table = gen_table_wide.iloc[:, first_hour:first_hour+24]
        gen_day_table.reset_index(inplace=True)
        gen_day_table = pd.merge(gen_data_table, gen_day_table, on='clave_ds')
        file_name = f"GEN_{files_name_suffix}_{day}"
        gen_day_table.to_csv(f"{out_folder}/{file_name}.csv", index=False, sep=";", decimal=",")

        dem_day_table = dem_table_wide.iloc[:, first_hour:first_hour+24]
        dem_day_table.reset_index(inplace=True)
        dem_day_table = pd.merge(dem_data_table, dem_day_table, on='clave_ds')
        file_name = f"DEM_{files_name_suffix}_{day}"     
        dem_day_table.to_csv(f"{out_folder}/{file_name}.csv", index=False, sep=";", decimal=",")


    return


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--data_folder', type=str, default='src/db_ds', help='data folder')
    parser.add_argument('--out_folder', type=str, default='files_ds', help='.csv output folder')
    parser.add_argument('--db_folder_name', type=str, default='2022', help='database folder name')
    parser.add_argument('--db_name', type=str, default='BD_balance_valorizado_2209_Data.accdb', help='data base name')
    parser.add_argument('--excel_name', type=str, default='BD_DS_2209.xlsm', help='data base name')

    args = parser.parse_args()

    # convert to dictionary
    params = vars(args)

    pypyodbc.lowercase = False
    db_path = f"{params['data_folder']}/{params['db_folder_name']}"
    db_name = params['db_name']
    files_name_suffix=db_name[-15:-11]
    out_path = f"{params['out_folder']}/{params['db_folder_name']}/{files_name_suffix}"
    if not os.path.exists(out_path):
        os.mkdir(out_path)

    db_name = params['db_name']
    excel_name = params['excel_name']
    files_name_suffix=db_name[-15:-11]

    time_start = timeit.default_timer()
    create_property_table(db_path, out_path, db_name, excel_name, files_name_suffix)
    time_end = timeit.default_timer()
    print(f"time run = {time_end - time_start}")
