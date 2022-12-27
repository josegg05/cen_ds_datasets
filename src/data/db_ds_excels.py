import argparse
import timeit
import os
from datetime import date

import numpy as np
import pandas as pd
import pypyodbc
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_template(path, excel_name):
    workbook = load_workbook(filename=f"{path}/{excel_name}")
    print(workbook.sheetnames)
    columns_names = ['Clave','Tipo','nombre_barra','1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24','Zona','Tag','Check','Checkeador nuevos retiros']

    sheet = workbook['Inyecciones Fact']
    #sheet["A1"] = "hello"
    #sheet["B1"] = "world!"
    #sheet.delete_cols(idx=1, amount=7)
    sheet.insert_cols(idx=7, amount=21)
    sheet['A2:AF2'] = columns_names
    

    workbook.save(filename="hello_world.xlsx")


def load_tables(data_path, files_date):
    gen_table = pd.read_csv(f'{data_path}/GENERATION_{files_date}.csv', sep=";", decimal=",")
    dem_table = pd.read_csv(f'{data_path}/DEMAND_{files_date}.csv', sep=";", decimal=",")
    return gen_table, dem_table


def create_excel(path, base_excel_name, out_path, out_file_name):
    workbook = load_workbook(filename=f"{path}/{base_excel_name}")  # , keep_vba=True)  # puede que esto preserve los macros
    workbook.save(filename=f"{out_path}/{out_file_name}")
    return


def insert_tables_to_excel(gen_table, dem_table, path, excel_name):
    workbook = load_workbook(filename=f"{path}/{excel_name}")
    var_type_list = ['gen', 'dem']
    
    for var_type in var_type_list:
        if var_type == 'gen':
            var_table = gen_table
            sheet = workbook['Inyecciones Fact']
            firts_excel_col = 1
            first_excel_row = 3
        elif var_type == 'dem':
            var_table = dem_table
            sheet = workbook['Retiros Fact']
            firts_excel_col = 1
            first_excel_row = 3

        last_table_col = var_table.shape[1]            
        last_table_row = var_table.shape[0]

        # Extract list of excel devices
        for value in sheet.iter_cols(min_row=first_excel_row, min_col=firts_excel_col, max_col=firts_excel_col, values_only=True):
            excel_devices = [str(i) for i in value]

        table_valid_devices = []
        table_invalid_devices = []
        excel_row_offset = 0
        for table_row in range(last_table_row):
            device_vars = list(var_table.iloc[table_row, :])
            if device_vars[0] in excel_devices:
                table_valid_devices.append(device_vars[0])
                for table_col in range(last_table_col):
                    sheet.cell(row=first_excel_row + excel_row_offset, column=firts_excel_col + table_col).value = device_vars[table_col]
                excel_row_offset+=1
            else:
                table_invalid_devices.append(device_vars[0])

        last_excel_row = first_excel_row + len(table_valid_devices)
        dev_with_no_data = [dev for dev in excel_devices if dev not in (table_valid_devices + table_invalid_devices)]
        print(f'# excel devices = {len(excel_devices)}; # table_valid_devices = {len(table_valid_devices)}')
        print(f'excel dev with no data = {len(excel_devices) - len(table_valid_devices)}: {dev_with_no_data}')
        print(f'table_invalid_devices = {len(table_invalid_devices)}: {table_invalid_devices}')

        sheet.delete_rows(idx=last_excel_row, amount=(len(excel_devices) - len(table_valid_devices)))

    workbook.save(filename=f"{path}/{excel_name}")
    return

def generate_days_excel(gen_table, dem_table, excel_path, out_path, files_date):
    base_excel_name = f"BD_DS_{files_date}_template.xlsm"
    attributes = ['clave_ds', 'tipo1', 'nombre_barra']   
    num_attributes = len(attributes)
    num_days = int((gen_table.shape[1] - num_attributes)/24)

    gen_base_table = gen_table[attributes]
    dem_base_table = dem_table[attributes]
    gen_data_table = gen_table.iloc[:, num_attributes+1:]
    dem_data_table = dem_table.iloc[:, num_attributes+1:]
    gen_zone_column = gen_table['Zona']
    dem_zone_column = dem_table['Zona']


    # num_days = 2  # BORRAR
    for day in range(num_days):
        # Create Excel
        print(f'********** Generating excel of day {day+1} ************')
        out_file_name = f"BD_DS.xlsx"
        out_path_day = f'{out_path}/BD_DS_dia_{day+1}'
        if not os.path.exists(out_path_day):
                os.mkdir(out_path_day)
        create_excel(excel_path, base_excel_name, out_path_day, out_file_name)

        first_hour = day*24
        gen_day_table = gen_data_table.iloc[:, first_hour:first_hour+24]
        #gen_day_table.reset_index(inplace=True)
        gen_day_table = pd.merge(gen_base_table, gen_day_table, left_index=True, right_index=True)  # , on='clave_ds')
        gen_day_table['Zona'] = gen_zone_column
        #file_name = f"GEN_{files_name_suffix}_{day}"
        #gen_day_table.to_csv(f"{out_folder}/{file_name}.csv", index=False, sep=";", decimal=",")

        dem_day_table = dem_data_table.iloc[:, first_hour:first_hour+24]
        #dem_day_table.reset_index(inplace=True)
        dem_day_table = pd.merge(dem_base_table, dem_day_table, left_index=True, right_index=True)  #, on='clave_ds')
        #dem_day_table['Zona'] = dem_zone_column
        #file_name = f"DEM_{files_name_suffix}_{day}"     
        #dem_day_table.to_csv(f"{out_folder}/{file_name}.csv", index=False, sep=";", decimal=",")

        insert_tables_to_excel(gen_day_table, dem_day_table, out_path_day, out_file_name)

    return


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--data_folder', type=str, default='src/data', help='data folder')
    parser.add_argument('--tables_folder_name', type=str, default='gen_dem_tables', help='database folder name')
    parser.add_argument('--excel_folder_name', type=str, default='model_ds', help='model and excel folder name')
    parser.add_argument('--data_year', type=str, default='2022', help='table year')
    parser.add_argument('--data_month', type=str, default='03', help='table month')
    parser.add_argument('--out_folder', type=str, default='model_ds', help='.csv output folder')


    args = parser.parse_args()

    # convert to dictionary
    params = vars(args)

    files_date=f'{params["data_year"][-2:]}{params["data_month"]}'
    data_path = f"{params['data_folder']}/{params['tables_folder_name']}/{params['data_year']}/{files_date}"
    excel_path = f"{params['data_folder']}/{params['excel_folder_name']}/{params['data_year']}/FPen_{files_date}_def"

    out_path = f"{params['data_folder']}/{params['out_folder']}"
    if not os.path.exists(out_path):
        os.mkdir(out_path)

    out_path = f"{out_path}/{params['data_year']}"
    if not os.path.exists(out_path):
        os.mkdir(out_path)

    out_path = f"{out_path}/FPen_{files_date}_def"
    if not os.path.exists(out_path):
        os.mkdir(out_path)

    # out_path = f"{out_path}/BD_DS_days"
    # if not os.path.exists(out_path):
    #     os.mkdir(out_path)
    

    time_start = timeit.default_timer()
    gen_table, dem_table = load_tables(data_path, files_date)
    generate_days_excel(gen_table, dem_table, excel_path, out_path, files_date)
    time_end = timeit.default_timer()
    print(f"time run = {time_end - time_start}")
